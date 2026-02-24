"""
LLM Extractor for HSA Receipt System
Uses vision-enabled LLM (Mistral Small 3) for direct image-to-JSON extraction
"""

import base64
import contextlib
import json
import logging
import re
import tempfile
from dataclasses import asdict, dataclass
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# Extraction constants
MIN_PAGE_TEXT_LENGTH = 50  # Skip pages with less usable text
MAX_FALLBACK_PAGES = 4  # Max pages to check in image-only fallback
MAX_PDF_PAGES = 5  # Max pages to process from PDFs

JSON_EXTRACTOR_SYSTEM_PROMPT = (
    "You are a JSON extractor. You ONLY output valid JSON objects. "
    "Never output markdown, explanations, or any text outside the JSON. "
    "Your response must start with { and end with }."
)


class Category(Enum):
    MEDICAL = "medical"
    DENTAL = "dental"
    VISION = "vision"
    PHARMACY = "pharmacy"
    UNKNOWN = "unknown"


class DocumentType(Enum):
    RECEIPT = "receipt"
    EOB = "eob"
    STATEMENT = "statement"
    CLAIM = "claim"
    PRESCRIPTION = "prescription"
    UNKNOWN = "unknown"


@dataclass
class ExtractedReceipt:
    """Structured receipt data extracted from document."""

    provider_name: str
    service_date: str | None  # YYYY-MM-DD
    service_type: str
    patient_name: str
    billed_amount: float
    insurance_paid: float
    patient_responsibility: float
    hsa_eligible: bool
    category: str
    document_type: str
    confidence_score: float
    notes: str
    raw_extraction: dict[str, Any]

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)

    def generate_filename(self, extension: str = "pdf") -> str:
        """Generate standardized filename."""
        date = self.service_date or datetime.now().strftime("%Y-%m-%d")
        provider = re.sub(r"[^\w\s-]", "", self.provider_name)[:30].strip().replace(" ", "_")
        service = re.sub(r"[^\w\s-]", "", self.service_type)[:20].strip().replace(" ", "_")
        amount = f"{self.patient_responsibility:.2f}"
        return f"{date}_{provider}_{service}_${amount}.{extension}"


@dataclass
class ExtractedClaim:
    """Single claim extracted from an EOB (one service line)."""

    service_date: str  # YYYY-MM-DD
    patient_name: str
    original_provider: str  # Provider who rendered service (e.g., "Stanford Health")
    service_type: str
    billed_amount: float
    insurance_paid: float
    patient_responsibility: float
    claim_number: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass
class MultiClaimExtraction:
    """Extraction result for multi-claim EOBs (e.g., Aetna EOB with multiple services)."""

    document_type: str  # "eob"
    payer_name: str  # "Aetna" (insurance company, not service provider)
    category: str  # "medical", "dental", "vision"
    confidence_score: float
    notes: str
    raw_extraction: dict[str, Any]
    claims: list[ExtractedClaim]
    statement_date: str = ""  # YYYY-MM-DD, used for unique filenames

    def to_dict(self) -> dict[str, Any]:
        result = asdict(self)
        result["claims"] = [c.to_dict() for c in self.claims]
        return result


EXTRACTION_PROMPT_TEMPLATE = """You are a medical receipt/EOB data extractor. Analyze this document image and extract structured information.

The family members are: {family_members}

Extract the following as a JSON object:
{{
  "provider_name": "Name of healthcare provider, pharmacy, or retailer",
  "service_date": "YYYY-MM-DD format or null if unclear",
  "service_type": "Brief description of service or product",
  "patient_name": "MUST be one of: {family_members} - match based on recipient/patient name in document",
  "eligible_subtotal": 0.00,
  "receipt_tax": 0.00,
  "receipt_taxable_amount": 0.00,
  "billed_amount": 0.00,
  "insurance_paid": 0.00,
  "patient_responsibility": 0.00,
  "hsa_eligible": true,
  "category": "medical|dental|vision|pharmacy",
  "document_type": "receipt|eob|statement|claim|prescription",
  "confidence_score": 0.95,
  "notes": "Any uncertainties or important details"
}}

CRITICAL - Recognize the store/provider from the image and apply these rules:

RETAIL STORES (Costco, CVS, Walgreens, Target, Walmart, Amazon):
- STRICT RULE: ONLY include items with VISIBLE "F" or "*" marker!
- The "F" marker appears in a dedicated COLUMN between price and "Dept" - look carefully!
- Items WITHOUT the "F" marker (like ZIPLOC, supplements) must be EXCLUDED
- Example from Costco receipt:
  * "SALONPAS 140    15.99 A  F  Dept" ← HAS "F" marker = INCLUDE
  * "ZIPLOC QUART    12.99 A     Dept" ← NO "F" marker = EXCLUDE
  * "NM COQ 140CT    37.99 A     Dept" ← NO "F" marker = EXCLUDE
- Rows starting with "SC" are discounts/coupons - IGNORE completely
- DO NOT assume eligibility based on product type - ONLY the "F" marker matters
- service_type: List EACH eligible item with quantity and unit price (e.g., "4x Salonpas @$15.99")
- DO NOT CALCULATE - just extract these raw values:
  * eligible_subtotal = sum of ONLY the FSA/HSA marked item prices (pre-tax)
  * receipt_tax = the TAX amount shown on receipt (e.g., "TAX 18.28" → 18.28)
  * receipt_taxable_amount = the taxable subtotal shown (e.g., "Taxable Amount 200.29" → 200.29)
- insurance_paid = 0 for retail purchases
- category = "pharmacy"
- If NO marked items found, set hsa_eligible=false and eligible_subtotal=0

HEALTHCARE EOBs (Sutter, Kaiser, Delta Dental, VSP, Anthem, Blue Cross):
- Look for "Patient Responsibility", "Member Pays", "Your Cost", "Amount Due"
- patient_responsibility = amount YOU owe after insurance
- insurance_paid = what the plan/insurance paid
- category based on provider type (medical/dental/vision)

General Rules:
- patient_name MUST be exactly one of: {family_members}
- Match the patient/recipient in the document to the closest family member name
- If unclear, default to the first family member ({default_patient})
- Respond with ONLY the JSON object, no other text
{provider_skill}"""


# Provider-specific extraction skills (activated based on filename/content hints)
PROVIDER_SKILLS = {
    "costco": """
COSTCO RECEIPT RULES:
- Look for "F" marker in column after price - ONLY these items are FSA-eligible
- Format: "ITEM_NAME    PRICE A  F  Dept" means FSA-eligible
- Format: "ITEM_NAME    PRICE A     Dept" means NOT eligible (no F)
- Rows starting with "SC" are discounts - IGNORE them
- provider_name: "Costco"
- eligible_subtotal: sum ONLY items with F marker
- receipt_tax: the TAX amount shown
- receipt_taxable_amount: the "Taxable Amount" shown
- document_type: "receipt"
- category: "pharmacy"

Example: If you see 4 lines of "SALONPAS 140  15.99 A F Dept" → eligible_subtotal = 63.96
""",
    "cvs": """
CVS-SPECIFIC RULES:

FIRST: Determine if this is a PRESCRIPTION or OTC purchase.

PRESCRIPTION (has Rx number, NDC, "AMOUNT DUE", "RETAIL PRICE"):
- provider_name: "CVS"
- document_type: "prescription"
- eligible_subtotal: 0 (IMPORTANT: must be 0 for prescriptions)
- patient_responsibility: the "AMOUNT DUE" on the receipt (copay after insurance)
- billed_amount: the "RETAIL PRICE" shown on the receipt
- insurance_paid: RETAIL PRICE minus AMOUNT DUE
- service_type: "Rx: MEDICATION_NAME STRENGTH" (e.g., "Rx: Tranexamic Acid 650mg")
- category: "pharmacy"
- hsa_eligible: true (all prescriptions are HSA eligible)

OTC PURCHASE (no Rx number, store receipt format):
- Look for FSA/HSA ELIGIBLE label on items
- Only include items marked FSA eligible
- document_type: "receipt"
- category: "pharmacy"

DATE FORMAT:
- CVS receipts use M/DD/YY format with 2-digit year
- "1/23/26" means January 23, {current_year} (NOT 2023)
- The current year is {current_year}, so "{current_year_short}" = {current_year}, etc.
- Use "DATE FILLED" as the service_date for prescriptions
- Format as YYYY-MM-DD in your output
""",
    "walgreens": """
WALGREENS-SPECIFIC RULES:
- Look for "FSA" or "HSA" markers next to eligible items
- Prescription copays are always HSA eligible
- For OTC items, only include if marked FSA/HSA eligible
- category = "pharmacy"
""",
    "amazon": """
AMAZON ORDER RULES:
- patient_name: Extract from "Ship to:" name
- service_date: Use "Order placed" date (not delivery date)
- service_type: Product name from the order
- DO NOT CALCULATE - read these values directly from the receipt:
  * eligible_subtotal: Use "Grand Total" amount (this already includes tax)
  * receipt_tax: 0 (tax is already in Grand Total)
  * receipt_taxable_amount: 0 (not needed - Grand Total is final)
- If "FSA or HSA eligible: $X.XX" line exists, that IS the amount to use
- provider_name: "Amazon"
- category: "pharmacy"
- document_type: "receipt"
""",
    "sutter": """
OUTPUT FORMAT: Return ONLY a JSON object. No markdown, no explanation, no text before or after.

This is a Sutter Health / PAMF statement that may contain MULTIPLE service lines/charges.

CRITICAL DISTINCTION - GUARANTOR vs PATIENT:
- The "Guarantor" (or "Responsible Party") is the person who PAYS the bill - NOT necessarily the patient
- The "Patient" is the person who RECEIVED the medical service
- ALWAYS use the PATIENT name (not guarantor) for patient_name
- If "Patient:" and "Guarantor:" are different people, use the Patient name
- If the document says "Patient: Charlie" and "Guarantor: Alice", the patient_name is "Charlie"

REQUIRED JSON STRUCTURE:
{"document_type":"statement","payer_name":"Sutter Health","category":"medical","confidence_score":0.95,"notes":"","claims":[{"service_date":"YYYY-MM-DD","patient_name":"PATIENT_NAME","original_provider":"DOCTOR_NAME","service_type":"SERVICE_DESCRIPTION","billed_amount":0.00,"insurance_paid":0.00,"patient_responsibility":0.00,"claim_number":""}]}

FIELD MAPPING:
- service_date: "Date of Service" for each line item (format as YYYY-MM-DD)
- patient_name: The PATIENT who received care (NOT the guarantor/responsible party)
- original_provider: Doctor/provider name (e.g., "Gilliam, Amy E, MD")
- service_type: Service description (e.g., "Office Visit", "Laboratory", "X-Ray")
- patient_responsibility: Amount owed for each service line
- insurance_paid: Insurance payment for each service line
- billed_amount: Original charge for each service line

RULES:
- Each service line/charge = one claim entry (do NOT combine them)
- If multiple dates of service appear, each is a separate claim
- If multiple service types appear on the same date, each is still a separate claim
- Look for itemized charges in the statement detail section
- Return ONLY the JSON object, nothing else
""",
    "aetna": """
OUTPUT FORMAT: Return ONLY a JSON object. No markdown, no explanation, no text before or after.

This is an Aetna EOB with MULTIPLE claims for MULTIPLE patients. Extract ALL claims from ALL patient sections.

REQUIRED JSON STRUCTURE:
{"document_type":"eob","payer_name":"Aetna","category":"medical","confidence_score":0.95,"claims":[{"service_date":"YYYY-MM-DD","patient_name":"NAME","original_provider":"PROVIDER","service_type":"SERVICE","billed_amount":0.00,"insurance_paid":0.00,"patient_responsibility":0.00,"claim_number":""}]}

CRITICAL - PATIENT NAME MAPPING:
- The EOB has sections like "Claim for NAME (self)", "Claim for NAME (spouse)", "Claim for NAME (son)"
- Map "(self)" or "(subscriber)" -> first family member
- Map "(spouse)" -> second family member
- Map "(son)" or "(daughter)" or "(dependent)" -> third family member
- EVERY claim section has a different patient — do NOT assign all claims to the same person
- You MUST extract claims from ALL patient sections in the document

FIELD MAPPING (from the claim detail tables, NOT the payment summary):
- service_date: The date in the "Service type and date" column (format as YYYY-MM-DD)
- original_provider: The "Provider:" line under each "Claim for..." header
- patient_responsibility: Column I "Your share C+D+E+H=I" — this is the total the patient owes
- insurance_paid: Column G "Plan's share"
- billed_amount: Column A "Amount billed"
- claim_number: The "Claim ID:" value

RULES:
- Extract from the DETAIL tables (with columns A through I), NOT the payment summary at the top
- Each claim detail table = one claim entry
- Return ONLY the JSON object, nothing else
""",
    "express_scripts": """
OUTPUT FORMAT: Return ONLY a JSON object. No markdown, no explanation, no text before or after.

This is an Express Scripts (PBM) Claims Summary with MULTIPLE prescription claims for one or more family members.

REQUIRED JSON STRUCTURE:
{"document_type":"eob","payer_name":"Express Scripts","category":"pharmacy","confidence_score":0.95,"notes":"","claims":[{"service_date":"YYYY-MM-DD","patient_name":"NAME","original_provider":"Express Scripts","service_type":"MEDICATION_NAME","billed_amount":0.00,"insurance_paid":0.00,"patient_responsibility":0.00,"claim_number":""}]}

PATIENT NAME MAPPING:
- Map patient names to one of the configured family members
- Look for "Member Name", "Patient", or name at the top of each claim section

FIELD MAPPING:
- service_date: Date the prescription was filled or shipped (format as YYYY-MM-DD)
- original_provider: Always "Express Scripts" (the PBM)
- service_type: Medication name (e.g., "Omeprazole 20mg", "Amoxicillin 500mg")
- patient_responsibility: "Your Cost", "You Pay", "Member Cost", "Copay" amount
- insurance_paid: "Plan Paid" or total minus your cost
- billed_amount: "Total Cost" or "Drug Cost"
- claim_number: Rx number if shown

RULES:
- Each prescription line = one claim entry
- If multiple family members appear, create separate claims for each
- All prescriptions are HSA-eligible (hsa_eligible = true)
- Return ONLY the JSON object, nothing else
""",
    "delta_dental": """
DELTA DENTAL-SPECIFIC RULES:
- This is a dental EOB
- Look for "Patient Pays" for patient_responsibility
- Look for "Benefit Paid" for insurance_paid
- category = "dental"
- document_type = "eob"
""",
    "vsp": """
VSP VISION-SPECIFIC RULES:
- This is a vision EOB or receipt
- Look for "Your Cost" or "Member Pays" for patient_responsibility
- category = "vision"
""",
    "stanford": """
STANFORD HEALTH CARE-SPECIFIC RULES:
- This is a hospital/medical statement from Stanford Health Care
- IMPORTANT: Look through ALL text for the key fields - they may be on different pages
- Look for "Patient Responsibility" for the amount the patient owes
- Look for "Balance Due" or "Amount Due" as confirmation
- Look for "Patient Deductible" - this often equals Patient Responsibility for HDHP plans
- "Service Date" is the date of medical service (not statement date)
- "Visit Type" describes the service (e.g., Outpatient, Inpatient)
- "Location" is usually "Stanford Hospital" or a clinic name
- Look for service descriptions like "Treatment/Observation Room"
- insurance_paid may be $0 if applied to deductible
- billed_amount is "Total Charges"
- category = "medical"
- document_type = "statement"
""",
}


def detect_provider_skill(filename: str, hints: list[str] | None = None) -> str | None:
    """Detect which provider skill to apply based on filename or hints.

    Args:
        filename: Name of the file being processed
        hints: Optional list of text hints (e.g., from OCR preview)

    Returns:
        Provider skill key if detected, None otherwise
    """
    text_to_check = filename.lower()
    if hints:
        text_to_check += " " + " ".join(h.lower() for h in hints)

    # Check for provider matches
    # NOTE: EOB routing to EOBs/{category}/ folder is planned for v0.3.0
    provider_patterns = {
        "costco": ["costco", "store 423", "store423"],  # Costco store numbers
        "cvs": ["cvs"],
        "walgreens": ["walgreens", "walgreen"],
        "amazon": ["amazon"],
        "express_scripts": [
            "express scripts",
            "express_scripts",
            "express_script",
            "express-scripts",
            "expressscripts",
            "esrx",
        ],
        "sutter": ["sutter", "pamf", "palo alto medical"],
        "aetna": ["aetna"],
        "delta_dental": ["delta dental", "deltadental"],
        "vsp": ["vsp", "vision service plan"],
        "stanford": ["stanford", "stanford health", "stanfordhealthcare"],
    }

    for skill_key, patterns in provider_patterns.items():
        for pattern in patterns:
            if pattern in text_to_check:
                return skill_key

    return None


def get_extraction_prompt(
    family_members: list[str] | None = None,
    provider_skill: str | None = None,
) -> str:
    """Generate extraction prompt with family member names and optional provider skill.

    Args:
        family_members: List of family member names
        provider_skill: Optional provider skill key (e.g., "costco", "cvs")

    Returns:
        Complete extraction prompt
    """
    family_members = family_members or ["Alice", "Bob", "Charlie"]

    skill_text = ""
    if provider_skill and provider_skill in PROVIDER_SKILLS:
        skill_text = PROVIDER_SKILLS[provider_skill]
        logger.info(f"Applying provider skill: {provider_skill}")

    # Inject dynamic values into provider skill text
    now = datetime.now()
    skill_text = skill_text.replace("{current_year}", str(now.year))
    skill_text = skill_text.replace("{current_year_short}", str(now.year % 100))

    return EXTRACTION_PROMPT_TEMPLATE.format(
        family_members=", ".join(family_members),
        default_patient=family_members[0],
        provider_skill=skill_text,
    )


class VisionExtractor:
    """Extract structured data from document images using vision-enabled LLM."""

    def __init__(
        self,
        api_base: str = "http://localhost:11434/v1",
        model: str = "mistral-small3",
        vision_model: str | None = None,
        eob_model: str | None = None,
        max_tokens: int = 2048,
        temperature: float = 0.1,
        family_members: list[str] | None = None,
    ):
        self.api_base = api_base.rstrip("/")
        self.model = model
        self.vision_model = vision_model or model  # Fallback to primary model
        self.eob_model = eob_model or model  # Larger model for complex EOBs
        self.max_tokens = max_tokens
        self.temperature = temperature
        self.family_members = family_members or ["Alice", "Bob", "Charlie"]
        self._client = None
        self._current_provider_skill = None  # Set per-file extraction

    def _init_client(self):
        if self._client is None:
            try:
                from openai import OpenAI

                self._client = OpenAI(
                    base_url=self.api_base,
                    api_key="ollama",  # Ollama doesn't need real key
                )
                logger.info(f"Vision LLM client initialized: {self.api_base}, model: {self.model}")
            except ImportError as err:
                raise ImportError("openai package not installed. Run: uv add openai") from err
        return self._client

    def _encode_image(self, image_path: Path) -> tuple[str, str]:
        """Encode image to base64 and determine MIME type."""
        suffix = image_path.suffix.lower()
        mime_types = {
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
            ".png": "image/png",
            ".gif": "image/gif",
            ".webp": "image/webp",
        }
        mime_type = mime_types.get(suffix, "image/jpeg")

        with open(image_path, "rb") as f:
            image_data = base64.b64encode(f.read()).decode("utf-8")

        return image_data, mime_type

    def _convert_pdf_to_images(self, pdf_path: Path, max_pages: int = MAX_PDF_PAGES) -> list[Path]:
        """Convert PDF pages to images for vision processing."""
        try:
            from pdf2image import convert_from_path
        except ImportError as err:
            raise ImportError("pdf2image not installed. Run: uv add pdf2image") from err

        images = convert_from_path(str(pdf_path), dpi=200, last_page=max_pages)
        image_paths = []

        for i, image in enumerate(images):
            temp_path = Path(tempfile.gettempdir()) / f"hsa_receipt_page_{i}.png"
            image.save(temp_path, "PNG")
            image_paths.append(temp_path)

        return image_paths

    def _decode_cid_text(self, text: str) -> str:
        """Decode CID-encoded text like (cid:84)(cid:104) to actual characters."""

        def decode_cid(match):
            try:
                cid = int(match.group(1))
                # CID values are typically ASCII codes
                if cid == 10:  # Newline - check first
                    return "\n"
                elif 32 <= cid <= 126:  # Printable ASCII (includes space at 32)
                    return chr(cid)
                else:
                    return ""
            except (ValueError, OverflowError):
                return ""

        return re.sub(r"\(cid:(\d+)\)", decode_cid, text)

    def _clean_extracted_text(self, text: str) -> str:
        """Clean extracted PDF text by removing garbage and decoding CID."""
        # First, decode CID-encoded text like (cid:84)(cid:104) -> "Th"
        text = self._decode_cid_text(text)

        # Remove QR code binary patterns (long strings of 0s and 1s)
        text = re.sub(r"\b[01]{10,}\b", "", text)

        # Remove hex patterns like 0X37B08973
        text = re.sub(r"\b0X[0-9A-Fa-f]+\b", "", text)

        # Remove excessive whitespace
        text = re.sub(r"\n\s*\n\s*\n+", "\n\n", text)
        text = re.sub(r"[ \t]+", " ", text)

        return text.strip()

    def _extract_text_with_pdfplumber(self, pdf_path: Path, max_pages: int = MAX_PDF_PAGES) -> str:
        """Extract text from PDF pages using pdfplumber."""
        try:
            import pdfplumber
        except ImportError as err:
            raise ImportError("pdfplumber not installed. Run: uv add pdfplumber") from err

        all_text = []
        try:
            with pdfplumber.open(pdf_path) as pdf:
                pages_to_process = min(len(pdf.pages), max_pages)
                for i, page in enumerate(pdf.pages[:pages_to_process]):
                    page_text = self._clean_extracted_text(page.extract_text() or "")
                    if len(page_text) > MIN_PAGE_TEXT_LENGTH:
                        all_text.append(f"=== PAGE {i + 1} ===\n{page_text}")

            combined = "\n\n".join(all_text)
            logger.info(
                f"Extracted {len(combined)} chars from {len(all_text)} PDF pages (of {pages_to_process} processed)"
            )
            return combined
        except (OSError, ValueError) as e:
            logger.warning(f"pdfplumber extraction failed ({type(e).__name__}): {e}")
            return ""
        except Exception as e:
            logger.error(f"Unexpected pdfplumber error: {type(e).__name__}: {e}")
            return ""

    def _extract_with_text_and_image(
        self, text_content: str, image_path: Path | None = None
    ) -> ExtractedReceipt:
        """Extract receipt data using text content and optional image."""
        client = self._init_client()
        prompt = self._get_prompt()

        # Build message content
        content = []

        # Add instruction about text content
        text_prompt = (
            prompt
            + f"""

DOCUMENT TEXT (extracted from all pages):
```
{text_content}
```

Analyze the text above to extract the JSON data. The text contains content from ALL pages of the document.
"""
        )
        content.append({"type": "text", "text": text_prompt})

        # Add image if available (for visual verification)
        if image_path and image_path.exists():
            image_data, mime_type = self._encode_image(image_path)
            content.append(
                {
                    "type": "image_url",
                    "image_url": {"url": f"data:{mime_type};base64,{image_data}"},
                }
            )

        try:
            # Use vision model when image is included, text model otherwise
            model = self.vision_model if (image_path and image_path.exists()) else self.model
            response = client.chat.completions.create(
                model=model,
                messages=[{"role": "user", "content": content}],
                max_tokens=self.max_tokens,
                temperature=self.temperature,
            )

            raw_response = response.choices[0].message.content
            parsed = self._parse_response(raw_response)
            return self._build_receipt(parsed)

        except Exception as e:
            logger.error(f"Text+image extraction failed: {e}")
            return self._fallback_extraction("text extraction")

    def _get_prompt(self) -> str:
        """Get the extraction prompt, including any active provider skill."""
        return get_extraction_prompt(
            family_members=self.family_members,
            provider_skill=self._current_provider_skill,
        )

    def _extract_xlsx_claims(self, file_path: Path) -> MultiClaimExtraction:
        """Extract claims directly from an Express Scripts xlsx file.

        No LLM needed - pure structured data parsing via openpyxl.

        Expected columns:
        A: Fullname With YOB  B: Prescription Number  C: Drug Name
        D: Drug Strength  E: Days Supply  F: Quantity  G: Pharmacy Name
        H: Date Of Service  I: Total cost  J: Total you paid
        """
        try:
            import openpyxl
        except ImportError as err:
            raise ImportError("openpyxl not installed. Run: uv add openpyxl") from err

        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active
        claims = []

        # Read header row to find columns dynamically
        headers = {}
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            if cell.value:
                headers[cell.value.strip().lower()] = cell.column - 1  # 0-indexed

        # Map expected column names
        col_map = {
            "name": headers.get("fullname with yob", 0),
            "rx_number": headers.get("prescription number", 1),
            "drug_name": headers.get("drug name", 2),
            "drug_strength": headers.get("drug strength", 3),
            "days_supply": headers.get("days supply", 4),
            "quantity": headers.get("quantity", 5),
            "pharmacy": headers.get("pharmacy name", 6),
            "date": headers.get("date of service", 7),
            "total_cost": headers.get("total cost", 8),
            "you_paid": headers.get("total you paid", 9),
        }

        for row in ws.iter_rows(min_row=2, values_only=True):
            # Skip empty rows and total rows
            if not row[col_map["name"]]:
                continue

            raw_name = str(row[col_map["name"]])
            # Strip year of birth suffix like "(2017)"
            patient_raw = re.sub(r"\s*\(\d{4}\)\s*$", "", raw_name).strip()
            patient = self._map_patient_name(patient_raw)

            drug_name = str(row[col_map["drug_name"]] or "Unknown")
            drug_strength = str(row[col_map["drug_strength"]] or "")
            service_type = f"{drug_name} {drug_strength}".strip()

            # Parse date (MM/DD/YYYY -> YYYY-MM-DD)
            raw_date = str(row[col_map["date"]] or "")
            try:
                parsed_date = datetime.strptime(raw_date, "%m/%d/%Y")
                service_date = parsed_date.strftime("%Y-%m-%d")
            except ValueError:
                service_date = raw_date

            total_cost = self._parse_amount(row[col_map["total_cost"]])
            you_paid = self._parse_amount(row[col_map["you_paid"]])
            days_supply = row[col_map["days_supply"]] or ""
            quantity = row[col_map["quantity"]] or ""
            rx_number = str(row[col_map["rx_number"]] or "")

            claims.append(
                ExtractedClaim(
                    service_date=service_date,
                    patient_name=patient,
                    original_provider=str(row[col_map["pharmacy"]] or "Express Scripts"),
                    service_type=service_type,
                    billed_amount=total_cost,
                    insurance_paid=max(0, total_cost - you_paid),
                    patient_responsibility=you_paid,
                    claim_number=rx_number,
                )
            )

            logger.info(
                f"Parsed xlsx claim: {patient} | {service_date} | "
                f"{service_type} | ${you_paid:.2f} (Rx#{rx_number}, "
                f"{days_supply}d supply, qty {quantity})"
            )

        wb.close()

        logger.info(f"Extracted {len(claims)} claims from xlsx (no LLM needed)")

        return MultiClaimExtraction(
            document_type="eob",
            payer_name="Express Scripts",
            category="pharmacy",
            confidence_score=1.0,  # Perfect - structured data, no LLM guessing
            notes="Parsed directly from Express Scripts xlsx export",
            raw_extraction={"source": "xlsx", "file": str(file_path)},
            claims=claims,
        )

    def _parse_aetna_eob(self, file_path: Path) -> MultiClaimExtraction | None:
        """Parse Aetna EOB deterministically from pdfplumber text.

        Extracts claims from the payment summary table on pages 1-2, plus
        claim details (claim IDs, service dates, service types) from the
        detail sections on subsequent pages.
        """
        # Read all pages — claim details span pages 2-7+ for large EOBs
        text = self._extract_text_with_pdfplumber(file_path, max_pages=20)
        if not text:
            return None

        # Strip page markers and repeated page headers so regex can match across pages
        text = re.sub(r"=== PAGE \d+ ===\n", "", text)
        text = re.sub(
            r"Statement date: .+? Page \d+ of \d+\n"
            r"Member: .+? Member ID: \S+\n"
            r"Group name: .+? Group #: \S+.*\n",
            "",
            text,
        )

        # --- Extract statement date ---
        stmt_date_match = re.search(r"Statement date:\s*(\w+ \d{1,2},\s*\d{4})", text)
        stmt_date = ""
        if stmt_date_match:
            try:
                dt = datetime.strptime(stmt_date_match.group(1), "%B %d, %Y")
                stmt_date = dt.strftime("%Y-%m-%d")
            except ValueError:
                pass

        # --- Parse payment summary table ---
        # Extract the block between "Your payment summary" and "Total:"
        summary_match = re.search(
            r"Your payment summary\s*\n(.+?)Total:\s*\$",
            text,
            re.DOTALL,
        )
        if not summary_match:
            logger.warning("Aetna parser: payment summary section not found")
            return None

        summary_block = summary_match.group(1)
        # Skip header lines (Plan's share / Your share / Patient Provider Amount...)
        header_end = re.search(r"Patient\s+Provider\s+Amount", summary_block)
        if header_end:
            summary_block = summary_block[header_end.end() :]
        # Also skip subheaders like "Sent to Send date Amount"
        subheader = re.search(r"Sent to\s+Send date\s+Amount\s*\n", summary_block)
        if subheader:
            summary_block = summary_block[subheader.end() :]

        # --- Build patient name mapping from claim detail headers ---
        # "Claim for Ming Hsun (self)" → {"Ming Hsun": "self"}
        # This lets us identify patient names in the payment summary
        name_role_map: dict[str, str] = {}
        for m in re.finditer(r"Claim for (.+?)\s*\((self|spouse|son|daughter|dependent)\)", text):
            name_role_map[m.group(1).strip()] = m.group(2).lower()
        # Sort by length descending so longer names match first
        known_names = sorted(name_role_map.keys(), key=len, reverse=True)

        # --- Parse payment summary lines ---
        role_pattern = re.compile(r"\((self|spouse|son|daughter|dependent)\)")
        dollar_pattern = re.compile(r"\$([\d,]+\.\d{2})")

        raw_lines = summary_block.strip().split("\n")
        # Pre-process: merge bare role markers like "(spouse)" into previous line
        merged_lines: list[str] = []
        for line in raw_lines:
            stripped = line.strip()
            if not stripped or stripped.startswith("Continued") or stripped.startswith("Page "):
                continue
            if "Patient" in stripped and "Provider" in stripped:
                continue
            if "Sent to" in stripped and "Send date" in stripped:
                continue
            if "Plan's share" in stripped and "Your share" in stripped:
                continue
            # Bare role marker (no dollar amounts) — splice into previous line
            bare_role = role_pattern.match(stripped)
            if bare_role and not dollar_pattern.search(stripped):
                if merged_lines:
                    merged_lines[-1] += " " + stripped
                continue
            # Line with dollar amounts — either a new claim or continuation
            if dollar_pattern.search(stripped):
                # Check if it starts with a known patient name or has a role marker
                has_role = role_pattern.search(stripped)
                starts_with_patient = any(stripped.startswith(n) for n in known_names)
                if has_role or starts_with_patient:
                    merged_lines.append(stripped)
                elif merged_lines:
                    # Continuation (e.g., "Inc." after "iRhythm Technologies,")
                    merged_lines[-1] += " " + stripped
            elif merged_lines:
                # Non-dollar continuation line (e.g., "Inc.")
                merged_lines[-1] += " " + stripped

        # Parse each merged line
        summary_claims = []
        for line in merged_lines:
            # Determine patient and role
            role = None
            patient_name_end = 0  # End of patient name text (before provider)

            # Try inline role marker first (role appears before any $)
            first_dollar_pos = line.find("$")
            if first_dollar_pos < 0:
                continue
            pre_dollar = line[:first_dollar_pos]
            role_match = role_pattern.search(pre_dollar)
            if role_match:
                role = role_match.group(1).lower()
                patient_name_end = role_match.end()
            else:
                # Role marker might be at end (bare merge) — check whole line
                role_match_end = role_pattern.search(line[first_dollar_pos:])
                if role_match_end:
                    role = role_match_end.group(1).lower()
                # Find patient name using known names from claim headers
                for name in known_names:
                    if pre_dollar.strip().startswith(name):
                        if role is None:
                            role = name_role_map[name]
                        patient_name_end = pre_dollar.index(name) + len(name)
                        break

            if role is None:
                continue

            patient = self._map_patient_name(f"({role})")

            # Provider is between patient name and first $
            provider = line[patient_name_end:first_dollar_pos].strip().rstrip(",.")

            # Extract all dollar amounts
            amounts = [float(m.group(1).replace(",", "")) for m in dollar_pattern.finditer(line)]
            if len(amounts) >= 2:
                plan_share = amounts[0]
                your_share = amounts[-1]
            elif len(amounts) == 1:
                plan_share = 0.0
                your_share = amounts[0]
            else:
                continue

            summary_claims.append(
                {
                    "patient_name": patient,
                    "original_provider": provider,
                    "insurance_paid": plan_share,
                    "patient_responsibility": your_share,
                }
            )

        if not summary_claims:
            logger.warning("Aetna parser: no claims found in payment summary")
            return None

        # --- Parse claim detail sections for dates, claim IDs, service types ---
        # Captures: patient role, provider, claim ID, and service block
        detail_pattern = re.compile(
            r"Claim for .+?\((self|spouse|son|daughter|dependent)\)\s*\n"
            r"Provider:\s*(.+?)\s*\(.*?\)\s*\n"
            r"Claim ID:\s*(\S+).*?"
            r"Service type and date\s*\n"
            r"A\s+B\s+C.*?\n"
            r"(.+?)(?=Claim for |Your Claim Remarks|Continued on next page|\Z)",
            re.DOTALL,
        )

        # Build lookup keyed by (patient, provider) to handle same provider for different patients
        # e.g., Matthew Lewis treats both Ming and Vanessa
        detail_lookup: dict[str, list[dict]] = {}
        for dm in detail_pattern.finditer(text):
            role = dm.group(1).lower()
            patient = self._map_patient_name(f"({role})")
            provider = dm.group(2).strip()
            claim_id = dm.group(3).strip()
            service_block = dm.group(4)

            # Extract service date: try "on\nM/D/YY", then bare "M/D/YY" on its own line
            date_match = re.search(r"on\s*\n\s*(\d{1,2}/\d{1,2}/\d{2,4})", service_block)
            if not date_match:
                # Bare date on its own line (no preceding "on")
                date_match = re.search(
                    r"^\s*(\d{1,2}/\d{1,2}/\d{2,4})\s*$", service_block, re.MULTILINE
                )
            svc_date = ""
            if date_match:
                raw_date = date_match.group(1)
                for fmt in ("%m/%d/%y", "%m/%d/%Y"):
                    try:
                        dt = datetime.strptime(raw_date, fmt)
                        svc_date = dt.strftime("%Y-%m-%d")
                        break
                    except ValueError:
                        continue

            # Extract service type (first line of service block, before amounts)
            svc_line = service_block.strip().split("\n")[0].strip()
            # Strip everything from first dollar amount or "on" + date
            svc_type = re.sub(r"\s+on\b.*$", "", svc_line).strip()
            svc_type = re.sub(r"\s+\d[\d,]*\.\d{2}\b.*$", "", svc_type).strip()

            # Extract billed amount from totals line (line starting with a dollar amount)
            billed_match = re.search(r"^(\d[\d,]*\.\d{2})\s+", service_block, re.MULTILINE)
            billed = float(billed_match.group(1).replace(",", "")) if billed_match else 0.0

            key = f"{patient}|{provider}"
            detail_lookup.setdefault(key, []).append(
                {
                    "claim_number": claim_id,
                    "service_date": svc_date,
                    "service_type": svc_type,
                    "billed_amount": billed,
                    "provider_full": provider,
                }
            )

        # --- Merge summary claims with detail info ---
        def _find_detail_key(patient: str, provider: str) -> str:
            """Find detail lookup key, with prefix fallback for truncated names."""
            exact = f"{patient}|{provider}"
            if exact in detail_lookup:
                return exact
            # Prefix match: "iRhythm Technologies" matches "iRhythm Technologies, Inc."
            for k in detail_lookup:
                if k.startswith(f"{patient}|") and k.split("|", 1)[1].startswith(provider):
                    return k
            return exact

        detail_use_count: dict[str, int] = {}
        claims = []
        for sc in summary_claims:
            patient = sc["patient_name"]
            provider = sc["original_provider"]
            key = _find_detail_key(patient, provider)
            idx = detail_use_count.get(key, 0)
            detail_use_count[key] = idx + 1

            # Look up detail by (patient, provider) — use idx-th entry for repeated providers
            details_list = detail_lookup.get(key, [])
            detail = details_list[idx] if idx < len(details_list) else {}

            claims.append(
                ExtractedClaim(
                    service_date=detail.get("service_date", stmt_date),
                    patient_name=sc["patient_name"],
                    original_provider=detail.get("provider_full") or provider,
                    service_type=detail.get("service_type", "Medical Service"),
                    billed_amount=detail.get("billed_amount", 0.0),
                    insurance_paid=sc["insurance_paid"],
                    patient_responsibility=sc["patient_responsibility"],
                    claim_number=detail.get("claim_number"),
                )
            )

        logger.info(f"Aetna parser: extracted {len(claims)} claims deterministically")
        return MultiClaimExtraction(
            document_type="eob",
            payer_name="Aetna",
            category="medical",
            confidence_score=0.95,
            notes="",
            raw_extraction={"source": "aetna_parser", "file": str(file_path)},
            claims=claims,
            statement_date=stmt_date,
        )

    def extract_eob(
        self, file_path: str | Path, provider_hint: str | None = None
    ) -> MultiClaimExtraction:
        """Extract multi-claim data from an EOB file.

        Args:
            file_path: Path to the EOB file (PDF or image)
            provider_hint: Optional hint for provider skill detection

        Returns:
            MultiClaimExtraction with list of claims
        """
        file_path = Path(file_path)

        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        # Detect provider skill
        hints = [provider_hint] if provider_hint else None
        self._current_provider_skill = detect_provider_skill(file_path.name, hints)
        if self._current_provider_skill:
            logger.info(f"Detected provider skill for EOB: {self._current_provider_skill}")

        try:
            # xlsx files: parse directly, no LLM needed
            suffix = file_path.suffix.lower()
            if suffix == ".xlsx":
                logger.info("Detected xlsx file - using direct spreadsheet parsing")
                return self._extract_xlsx_claims(file_path)

            # Aetna EOBs: use deterministic parser (no LLM needed)
            if self._current_provider_skill == "aetna" and suffix == ".pdf":
                result = self._parse_aetna_eob(file_path)
                if result and result.claims:
                    return result
                logger.warning("Aetna parser returned no claims, falling back to LLM")

            # Extract text from PDF - prefer text-based extraction
            text_content = ""
            if suffix == ".pdf":
                text_content = self._extract_text_with_pdfplumber(file_path)

            client = self._init_client()
            skill_text = PROVIDER_SKILLS.get(self._current_provider_skill, "")

            if text_content:
                # Text-based extraction (works for most EOBs/statements)
                user_prompt = f"""Extract claims from this document.

{skill_text}

Family members: {", ".join(self.family_members)}

DOCUMENT TEXT:
{text_content[:16000]}

Output JSON only, starting with {{ and ending with }}:"""

                eob_model = self.eob_model
                if eob_model != self.model:
                    logger.info(f"Using EOB model: {eob_model}")
                response = client.chat.completions.create(
                    model=eob_model,
                    messages=[
                        {"role": "system", "content": JSON_EXTRACTOR_SYSTEM_PROMPT},
                        {"role": "user", "content": user_prompt},
                    ],
                    max_tokens=8192,
                    temperature=self.temperature,
                )
            else:
                # Image-based fallback (for scanned/image PDFs like Express Scripts)
                logger.info("No text extracted, using vision-based multi-claim extraction")

                image_paths = []
                if suffix == ".pdf":
                    image_paths = self._convert_pdf_to_images(file_path, max_pages=MAX_PDF_PAGES)
                elif suffix in {".png", ".jpg", ".jpeg", ".gif", ".webp"}:
                    image_paths = [file_path]

                if not image_paths:
                    logger.error("No text or images available for extraction")
                    return self._build_multi_claim_extraction({})

                prompt_text = f"""Extract claims from this document image.

{skill_text}

Family members: {", ".join(self.family_members)}

Output JSON only, starting with {{ and ending with }}:"""

                # Build vision content with all page images
                content = [{"type": "text", "text": prompt_text}]
                for img_path in image_paths:
                    image_data, mime_type = self._encode_image(img_path)
                    content.append(
                        {
                            "type": "image_url",
                            "image_url": {"url": f"data:{mime_type};base64,{image_data}"},
                        }
                    )

                try:
                    logger.info(f"Using vision model: {self.vision_model}")
                    response = client.chat.completions.create(
                        model=self.vision_model,
                        messages=[
                            {"role": "system", "content": JSON_EXTRACTOR_SYSTEM_PROMPT},
                            {"role": "user", "content": content},
                        ],
                        max_tokens=4096,
                        temperature=self.temperature,
                    )
                finally:
                    # Cleanup temp images (only if we created them from PDF)
                    if suffix == ".pdf":
                        for img_path in image_paths:
                            with contextlib.suppress(OSError):
                                img_path.unlink()

            raw_response = response.choices[0].message.content
            parsed = self._parse_response(raw_response)

            return self._build_multi_claim_extraction(parsed)

        finally:
            self._current_provider_skill = None

    def _build_multi_claim_extraction(self, parsed: dict[str, Any]) -> MultiClaimExtraction:
        """Build MultiClaimExtraction from parsed JSON.

        Handles two JSON formats:
        1. Expected format with "claims" array
        2. Alternative format with "summary.payment_summary" or "services" arrays
        """
        claims = []

        # Try expected format first
        raw_claims = parsed.get("claims", [])

        # Fallback: try alternative format from payment_summary
        # Check both top-level and nested in summary
        if not raw_claims:
            payment_summary = parsed.get("payment_summary", [])
            if not payment_summary and "summary" in parsed:
                payment_summary = parsed.get("summary", {}).get("payment_summary", [])
            statement_date = parsed.get("statement_info", {}).get("statement_date", "")
            if statement_date == "Not specified":
                statement_date = ""

            for item in payment_summary:
                # Skip if item is not a dict (sometimes LLM returns strings)
                if not isinstance(item, dict):
                    continue

                amount = self._parse_amount(item.get("your_share"))
                plan_paid = self._parse_amount(item.get("plan_share"))

                raw_claims.append(
                    {
                        "service_date": statement_date,
                        "patient_name": self._map_patient_name(item.get("patient", "")),
                        "original_provider": item.get("provider", "Unknown"),
                        "service_type": "Medical Service",
                        "billed_amount": amount,
                        "insurance_paid": plan_paid,
                        "patient_responsibility": amount,
                    }
                )

        # Fallback: try services array for more detail
        if not raw_claims and "services" in parsed:
            for service_group in parsed.get("services", []):
                patient = self._map_patient_name(service_group.get("patient", ""))

                for detail in service_group.get("service_details", []):
                    cost = self._parse_amount(detail.get("your_cost"))
                    billed = self._parse_amount(detail.get("amount_billed"))

                    raw_claims.append(
                        {
                            "service_date": detail.get("date", ""),
                            "patient_name": patient,
                            "original_provider": detail.get("provider", "Unknown"),
                            "service_type": detail.get("service", "Medical Service"),
                            "billed_amount": billed,
                            "insurance_paid": max(0, billed - cost),
                            "patient_responsibility": cost,
                        }
                    )

        # LLM responses use varying field names; normalize to canonical names
        _field_aliases = {
            "patient": "patient_name",
            "provider": "original_provider",
            "your_share": "patient_responsibility",
            "plan_amount": "insurance_paid",
            "member_amount": "billed_amount",
            "claim_id": "claim_number",
        }

        for raw_claim in raw_claims:
            for alias, canonical in _field_aliases.items():
                if alias in raw_claim and canonical not in raw_claim:
                    raw_claim[canonical] = raw_claim[alias]

            # Ensure patient name is mapped
            patient = raw_claim.get("patient_name") or "Unknown"
            if patient not in self.family_members:
                patient = self._map_patient_name(patient)

            claim = ExtractedClaim(
                service_date=raw_claim.get("service_date") or "",
                patient_name=patient,
                original_provider=raw_claim.get("original_provider") or "Unknown",
                service_type=raw_claim.get("service_type") or "Unknown Service",
                billed_amount=self._parse_amount(raw_claim.get("billed_amount")),
                insurance_paid=self._parse_amount(raw_claim.get("insurance_paid")),
                patient_responsibility=self._parse_amount(raw_claim.get("patient_responsibility")),
                claim_number=raw_claim.get("claim_number"),
            )
            claims.append(claim)

        # Extract payer name - check multiple possible locations
        payer_name = parsed.get("payer_name") or "Unknown"
        if payer_name == "Unknown":
            # Try insurance_info.insurer or fall back to detected provider skill
            insurer = parsed.get("insurance_info", {}).get("insurer", "")
            if "aetna" in insurer.lower() or self._current_provider_skill == "aetna":
                payer_name = "Aetna"

        # Extract statement_date from LLM response (used for unique filenames)
        stmt_date = parsed.get("statement_info", {}).get("statement_date", "")
        if stmt_date == "Not specified":
            stmt_date = ""

        return MultiClaimExtraction(
            document_type=parsed.get("document_type") or "eob",
            payer_name=payer_name,
            category=parsed.get("category") or "medical",
            confidence_score=float(parsed.get("confidence_score") or 0.8),
            notes=parsed.get("notes") or "",
            raw_extraction=parsed,
            claims=claims,
            statement_date=stmt_date,
        )

    def _map_patient_name(self, raw_name: str) -> str:
        """Map raw patient names from EOB to family member names.

        Uses positional mapping for generic terms:
        - Index 0: primary account holder ("self", "subscriber")
        - Index 1: spouse ("spouse", "wife", "husband")
        - Index 2+: dependents ("son", "daughter", "child", "dependent")
        """
        raw_lower = raw_name.lower()

        # First, check for exact family member name match
        for name in self.family_members:
            if name.lower() in raw_lower:
                return name

        # Map generic terms to family members by position
        if len(self.family_members) >= 1 and any(
            kw in raw_lower for kw in ["self", "subscriber", "primary"]
        ):
            return self.family_members[0]

        if len(self.family_members) >= 2 and any(
            kw in raw_lower for kw in ["spouse", "wife", "husband"]
        ):
            return self.family_members[1]

        if len(self.family_members) >= 3 and any(
            kw in raw_lower for kw in ["son", "daughter", "child", "dependent"]
        ):
            return self.family_members[2]

        # Default to first family member
        return self.family_members[0] if self.family_members else "Unknown"

    def _parse_amount(self, value: Any) -> float:
        """Parse an amount value that might be a string with $ or a number."""
        if value is None:
            return 0.0
        if isinstance(value, int | float):
            return float(value)
        if not isinstance(value, str):
            return 0.0
        # Remove $ and , then convert
        cleaned = value.replace("$", "").replace(",", "").strip()
        if not cleaned:
            return 0.0
        try:
            return float(cleaned)
        except ValueError:
            return 0.0

    def extract_from_image(self, image_path: Path) -> ExtractedReceipt:
        """Extract receipt data from a single image using vision model."""
        client = self._init_client()
        image_data, mime_type = self._encode_image(image_path)
        prompt = self._get_prompt()

        try:
            response = client.chat.completions.create(
                model=self.vision_model,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": prompt},
                            {
                                "type": "image_url",
                                "image_url": {"url": f"data:{mime_type};base64,{image_data}"},
                            },
                        ],
                    }
                ],
                max_tokens=self.max_tokens,
                temperature=self.temperature,
            )

            raw_response = response.choices[0].message.content
            parsed = self._parse_response(raw_response)
            return self._build_receipt(parsed)

        except Exception as e:
            logger.error(f"Vision extraction failed: {e}")
            return self._fallback_extraction(str(image_path))

    def extract_from_pdf(self, pdf_path: Path) -> ExtractedReceipt:
        """Extract receipt data from PDF using pdfplumber text + first page image."""
        # First, try to extract text from all pages using pdfplumber
        text_content = self._extract_text_with_pdfplumber(pdf_path)

        # Convert first page to image for visual context
        image_paths = self._convert_pdf_to_images(pdf_path)

        try:
            if text_content:
                # Use text from all pages + first page image
                first_image = image_paths[0] if image_paths else None
                result = self._extract_with_text_and_image(text_content, first_image)

                # If result looks incomplete (zero amount), try image-only on key pages
                if result.patient_responsibility == 0 and len(image_paths) > 1:
                    logger.info(
                        "Zero amount from text extraction, trying image-only on key pages..."
                    )
                    for img_path in image_paths[:MAX_FALLBACK_PAGES]:
                        alt_result = self.extract_from_image(img_path)
                        if alt_result.patient_responsibility > 0:
                            # Replace with image-only result if it found a valid amount
                            if alt_result.confidence_score >= result.confidence_score:
                                result = alt_result
                            break

                return result

            elif image_paths:
                # Fallback: no text extracted, use image-only approach
                logger.warning("No text extracted, falling back to image-only")
                result = self.extract_from_image(image_paths[0])

                # Check other pages if first page gave low confidence
                if len(image_paths) > 1 and result.confidence_score < 0.7:
                    for img_path in image_paths[1:]:
                        alt_result = self.extract_from_image(img_path)
                        if alt_result.confidence_score > result.confidence_score:
                            result = alt_result
                            break

                return result
            else:
                return self._fallback_extraction(str(pdf_path))

        finally:
            # Cleanup temp images
            for img_path in image_paths:
                with contextlib.suppress(OSError):
                    img_path.unlink()

    def extract(self, file_path: str | Path, provider_hint: str | None = None) -> ExtractedReceipt:
        """Extract receipt data from file (image or PDF).

        Args:
            file_path: Path to the receipt file
            provider_hint: Optional hint for provider skill detection (e.g., "costco")
        """
        file_path = Path(file_path)

        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        # Detect provider skill from filename (or explicit hint)
        hints = [provider_hint] if provider_hint else None
        self._current_provider_skill = detect_provider_skill(file_path.name, hints)
        if self._current_provider_skill:
            logger.info(f"Detected provider skill: {self._current_provider_skill}")

        suffix = file_path.suffix.lower()

        try:
            if suffix == ".pdf":
                return self.extract_from_pdf(file_path)

            # Standard image formats - process directly
            if suffix in {".png", ".jpg", ".jpeg", ".gif", ".webp"}:
                return self.extract_from_image(file_path)

            # Formats requiring conversion to PNG
            if suffix in {".tiff", ".bmp", ".heic", ".heif"}:
                return self._extract_with_conversion(file_path, suffix)

            raise ValueError(f"Unsupported file type: {suffix}")
        finally:
            # Reset provider skill after extraction
            self._current_provider_skill = None

    def _extract_with_conversion(self, file_path: Path, suffix: str) -> ExtractedReceipt:
        """Extract from image formats that require conversion to PNG first."""
        if suffix in {".heic", ".heif"}:
            try:
                import pillow_heif

                pillow_heif.register_heif_opener()
            except ImportError as err:
                raise ImportError("pillow-heif not installed. Run: uv add pillow-heif") from err
            logger.info("Converting HEIC to PNG for processing")

        from PIL import Image

        img = Image.open(file_path)
        temp_path = Path(tempfile.gettempdir()) / "hsa_receipt_converted.png"
        img.save(temp_path, "PNG")
        try:
            return self.extract_from_image(temp_path)
        finally:
            temp_path.unlink(missing_ok=True)

    def _parse_response(self, response: str) -> dict[str, Any]:
        """Parse LLM response to extract JSON."""
        response = response.strip()

        # Strip markdown code blocks
        if response.startswith("```json"):
            response = response[7:]
        if response.startswith("```"):
            response = response[3:]
        if response.endswith("```"):
            response = response[:-3]
        response = response.strip()

        try:
            parsed = json.loads(response)
            # Handle case where LLM returns array instead of object
            if isinstance(parsed, list) and len(parsed) > 0:
                parsed = parsed[0]
            return parsed if isinstance(parsed, dict) else {}
        except json.JSONDecodeError:
            # Try to find JSON object with nested structure (for multi-claim)
            # Look for outermost { ... } allowing nested braces
            brace_count = 0
            start = -1
            for i, c in enumerate(response):
                if c == "{":
                    if brace_count == 0:
                        start = i
                    brace_count += 1
                elif c == "}":
                    brace_count -= 1
                    if brace_count == 0 and start != -1:
                        try:
                            return json.loads(response[start : i + 1])
                        except json.JSONDecodeError:
                            pass
                        start = -1

            logger.warning("Could not parse LLM response as JSON")
            logger.debug(f"Raw LLM response (first 500 chars): {response[:500]}")
            return {}

    def _build_receipt(self, parsed: dict[str, Any]) -> ExtractedReceipt:
        """Build ExtractedReceipt from parsed JSON, calculating tax in Python."""
        # Extract raw values
        eligible_subtotal = float(parsed.get("eligible_subtotal") or 0)
        receipt_tax = float(parsed.get("receipt_tax") or 0)
        receipt_taxable_amount = float(parsed.get("receipt_taxable_amount") or 0)
        insurance_paid = float(parsed.get("insurance_paid") or 0)

        # Calculate tax on eligible items (Python does the math, not LLM)
        tax_on_eligible = 0.0
        tax_rate = 0.0
        if eligible_subtotal > 0 and receipt_taxable_amount > 0 and receipt_tax > 0:
            tax_rate = receipt_tax / receipt_taxable_amount
            tax_on_eligible = round(eligible_subtotal * tax_rate, 2)

        # For retail (eligible_subtotal > 0): patient_responsibility = eligible items + tax
        # For prescriptions/EOBs (eligible_subtotal = 0): use extracted values directly
        document_type = parsed.get("document_type") or "unknown"
        if document_type in ("receipt", "prescription") and eligible_subtotal > 0:
            patient_responsibility = eligible_subtotal + tax_on_eligible
            billed_amount = eligible_subtotal
        else:
            # EOB or other - use extracted values
            patient_responsibility = float(
                parsed.get("patient_responsibility") or eligible_subtotal
            )
            billed_amount = float(parsed.get("billed_amount") or eligible_subtotal)

        # Build notes with tax calculation if applicable
        notes = parsed.get("notes") or ""
        if tax_on_eligible > 0:
            tax_note = (
                f"Tax rate {tax_rate * 100:.3f}%, tax on eligible items: ${tax_on_eligible:.2f}"
            )
            notes = f"{tax_note}. {notes}" if notes else tax_note

        # Handle service_type as string (LLM sometimes returns a list)
        service_type = parsed.get("service_type") or "Unknown Service"
        if isinstance(service_type, list):
            service_type = ", ".join(str(s) for s in service_type)

        return ExtractedReceipt(
            provider_name=parsed.get("provider_name") or "Unknown",
            service_date=parsed.get("service_date"),
            service_type=service_type,
            patient_name=parsed.get("patient_name") or "Unknown",
            billed_amount=billed_amount,
            insurance_paid=insurance_paid,
            patient_responsibility=patient_responsibility,
            hsa_eligible=bool(parsed.get("hsa_eligible", True)),
            category=parsed.get("category") or "unknown",
            document_type=document_type,
            confidence_score=float(parsed.get("confidence_score") or 0.5),
            notes=notes,
            raw_extraction=parsed,
        )

    def _fallback_extraction(self, source: str) -> ExtractedReceipt:
        """Fallback when vision extraction fails."""
        return ExtractedReceipt(
            provider_name="Unknown (Extraction Failed)",
            service_date=None,
            service_type="Unknown (Needs Manual Review)",
            patient_name="Unknown",
            billed_amount=0,
            insurance_paid=0,
            patient_responsibility=0,
            hsa_eligible=True,
            category="unknown",
            document_type="unknown",
            confidence_score=0.1,
            notes=f"Vision extraction failed - manual review required. Source: {source}",
            raw_extraction={},
        )


class MockVisionExtractor(VisionExtractor):
    """Mock extractor for testing without running LLM."""

    def extract(self, file_path: str | Path) -> ExtractedReceipt:
        return ExtractedReceipt(
            provider_name="Mock Provider",
            service_date="2026-01-15",
            service_type="Test Service",
            patient_name="Test Patient",
            billed_amount=100.00,
            insurance_paid=80.00,
            patient_responsibility=20.00,
            hsa_eligible=True,
            category="medical",
            document_type="receipt",
            confidence_score=0.95,
            notes="Mock extraction for testing",
            raw_extraction={"mock": True},
        )


def get_extractor(
    use_mock: bool = False,
    api_base: str = "http://localhost:11434/v1",
    model: str = "mistral-small3",
    vision_model: str | None = None,
    **kwargs,
) -> VisionExtractor:
    """Factory function to get appropriate extractor."""
    if use_mock:
        return MockVisionExtractor()
    return VisionExtractor(api_base=api_base, model=model, vision_model=vision_model, **kwargs)


if __name__ == "__main__":
    import sys

    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        use_mock = "--mock" in sys.argv

        extractor = get_extractor(use_mock=use_mock)
        result = extractor.extract(file_path)

        print("Extracted Data:")
        print(json.dumps(result.to_dict(), indent=2, default=str))
        print(f"\nGenerated filename: {result.generate_filename()}")
    else:
        print("Usage: python llm_extractor.py <image_or_pdf_path> [--mock]")
